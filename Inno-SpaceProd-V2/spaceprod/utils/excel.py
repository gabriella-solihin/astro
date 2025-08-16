"""
Excel utils for working with .xlsx files
"""


import datetime as dt
import os
import re
import string
from itertools import islice
from typing import Tuple

import openpyxl
import pandas as pd
from openpyxl.workbook.workbook import Workbook

from inno_utils.loggers.log import log


class ExcelWriter:
    def __init__(self, workbook_path: str):
        self.opx = openpyxl

        self.path = workbook_path

        if not os.path.exists(self.path):
            log.info(f"Workbook does not exist, creating: {self.path}")
            self._create_blank_workbook(self.path)

        self.wb = self.opx.load_workbook(workbook_path)
        self.writer = pd.ExcelWriter(workbook_path, engine="openpyxl")
        self.writer.book = self.wb

    def done(self):
        """
        Always run this at the end when you are done to save your changes
        :return:
        """
        self.wb.save(self.path)
        self.writer.book.save(self.path)
        self.writer.save()
        self.writer.book.close()
        self.writer.close()
        self.wb.close()

    def paste_table(
        self,
        pdf: pd.DataFrame,
        sheet_name: str = "Article Summary",
        startrow: int = None,
        truncate_sheet: bool = False,
        **to_excel_kwargs,
    ) -> None:
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Args:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          pdf : dataframe to save to workbook

          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')

          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...

          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file

          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """

        log.info("Start populating Excel template")

        # ignore [engine] parameter if it was passed
        if "engine" in to_excel_kwargs:
            to_excel_kwargs.pop("engine")

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in self.writer.book.sheetnames:
            startrow = self.writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in self.writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = self.writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            self.writer.book.remove(self.writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            self.writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        self.writer.sheets = {ws.title: ws for ws in self.writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        pdf.to_excel(
            self.writer, sheet_name, startrow=startrow, index=False, **to_excel_kwargs
        )

        log.info("Done populating Excel template")

    def create_named_range(self, name: str, range_address: str) -> None:
        """
        Creates a new named range in the excel workbook.
        If a range with such name already exists it replaces it.
        :param workbook_path: full path to the workbook
        :param name: Range name, e.g.: 'data_main'
        :param range_address:  Range full address, e.g.: 'Sheet!$A$1:$A$5'
        :return: None
        """
        opx = self.opx
        new_range = opx.defined_name.DefinedName(name, attr_text=range_address)
        all_names = [x.name for x in self.wb.defined_names.__dict__["definedName"]]
        if name in all_names:
            self.wb.defined_names.delete("data_main")

            self.wb.defined_names.append(new_range)

    def change_cell_formatting(
        self, formatting: str, sheet_name: str, range_address: str
    ):

        ws = self.wb[sheet_name]
        range = ws[range_address]

        for i, v in enumerate(range):
            range[i].number_format = formatting

    @staticmethod
    def _create_blank_workbook(path):
        opx = openpyxl
        wb = opx.Workbook()
        wb.save(path)
        wb.close()


def col2num(col: str):
    """ converts a column letter to a number """
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord("A")) + 1
    return num


def get_row_col(ref: str) -> Tuple[int, int]:
    """
    attempts to find the row number and column number based on a given
    Excel cell reference

    Parameters
    ----------
    ref : str
        Excel cell reference (e.g. A1)

    Returns
    -------
    a tuple or row number and column number
    """
    try:
        row_find = re.findall(r"\d+", ref)
        assert len(row_find) == 1
        row = int(row_find[0])

        col_find = re.findall(r"[a-zA-Z]+", ref)
        assert len(col_find) == 1
        col = int(col2num(col_find[0]))
        assert row > 0
        assert col > 0

    except:

        msg = f"""
        Invalid cell reference: '{ref}', please check.
        Must be a single cell reference, e.g.: 'A1' or 'AB20'
        """
        raise Exception(msg)

    return row, col


def convert_string_dates_in_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensures that if a column has a date in it that was read as string, it converts it
    to dt.datetime
    Must be of yyyy-mm-dd format!
    :param df: a dataframe
    :return: a dataframe with dates in column names converted
    """

    cols = list(df.columns)
    new_cols = []
    for c in cols:

        if not isinstance(c, str):
            new_cols.append(c)
            continue

        try:
            new_cols.append(dt.datetime.strptime(c, "%Y-%m-%d"))
        except ValueError:
            new_cols.append(c)

    df.columns = new_cols

    return df


def read_excel_data(wb: Workbook, sheet_name: str, top_left_cell: str) -> pd.DataFrame:
    """
    reads Excel tabular array of data into a pd.DataFrame
    Parameters
    ----------

    wb : Workbook
         an openpyxl openpyxl.workbook.workbook.Workbook object

    sheet_name : str
        sheet name to read

    top_left_cell : str
        Excel cell reference that represents the top left corner of the
        tabular array to read (including the header)

    Returns
    -------
    dataset as pd.DataFrame
    """

    sheet_vals = wb[sheet_name].values
    sheet_vals = list(sheet_vals)
    sheet_vals = (islice(r, 0, None) for r in sheet_vals)
    df = pd.DataFrame(data=sheet_vals)

    first_row, first_col = get_row_col(top_left_cell)

    col_names = df.iloc[(first_row - 1), (first_col - 1) :].tolist()

    df_result = df.iloc[first_row:, (first_col - 1) :].reset_index(drop=True)
    df_result.columns = col_names

    df_result = df_result.infer_objects()

    df_result = convert_string_dates_in_columns(df_result)

    return df_result
